# -*- coding: utf-8 -*-
"""스마트오더 라벨 생성 핵심 엔진 (채널: CU / GS / E24)

설계 원칙
 - 센터 매핑은 '센터명' 기반(전 채널 공통 center.json). 센터코드를 안 쓰는 채널(E24)이 있어 이름으로 통일
 - 상품 매핑은 '상품코드' 기반: 상품코드 → 대표(한익스상품명) + 구분(대월/프리미엄)
 - 마스터는 JSON 파일. 업로드 중 미매핑 항목은 화면에서 즉시 등록 → 재사용
 - 부착양식(A4출력)은 기존 xlsx와 픽셀 단위로 동일한 스타일로 새로 생성(정적값·정확한 라벨 수)
"""
import os, json, io, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
# 마스터 저장 경로: 배포/exe 시 쓰기 가능한 폴더로 분리(SMARTORDER_MASTERS), 없으면 앱 폴더 masters/.
MASTER_DIR = os.environ.get("SMARTORDER_MASTERS") or os.path.join(BASE, "masters")
# seed(동봉 기본 마스터) 위치: exe에서는 번들 경로(SMARTORDER_SEED)로 지정.
SEED_DIR = os.environ.get("SMARTORDER_SEED") or os.path.join(BASE, "masters")

def ensure_masters_seeded():
    """MASTER_DIR가 비어 있으면 코드 동봉 seed(masters/)를 복사한다(최초 배포용)."""
    if os.path.abspath(MASTER_DIR) == os.path.abspath(SEED_DIR):
        return
    os.makedirs(MASTER_DIR, exist_ok=True)
    import shutil
    for fn in os.listdir(SEED_DIR):
        if not fn.endswith(".json"):
            continue
        dst = os.path.join(MASTER_DIR, fn)
        if not os.path.exists(dst):
            shutil.copy2(os.path.join(SEED_DIR, fn), dst)

# ----------------------------------------------------------------------------
# 채널 정의
# ----------------------------------------------------------------------------
CENTER_MASTER = "center.json"   # 전 채널 공통: {센터명: 이고센터}. 센터코드를 안 쓰는 채널(E24)이 있어 이름 키로 통일

CHANNELS = {
    "cu": {
        "name": "CU",
        "tag": "CU",
        # {jm} = 대월직매장 / 프리미엄직매장 (상품 구분에 따라 라벨별로 결정)
        "title_tpl": "{jm}\n(BGF리테일 CU스마트오더)",
        "center_label": "CU센터",
        "hub_col": "배송센터",       # 라벨출력(폼텍) 시트에 덧붙일 이고센터 컬럼명
        "label_product": "raw",     # 라벨 상품명 = 화주상품명(원본). 표지는 대표(한익스상품명)
        "col_width": {"A": 39.0, "B": 99.9, "C": 41.7, "D": 8.7},
        "row_h": 62.4,
        "lead_spacer": True,
        # RAW 헤더 후보 (부분일치)
        "raw": {
            "sheet_hint": ["발주확인", "스티커", "lowdata", "Sheet"],
            "header_scan_rows": 3,
            "center_code": ["센터 코드", "센터코드"],
            "center_name": ["센터명", "센터 명"],
            "store_code":  ["센터 코드", "센터코드"],
            "store_name":  ["센터명"],
            "prod_code":   ["상품 코드", "상품코드"],
            "prod_name":   ["상품명"],
            # CU 수량 = 발주단위수량(라벨 1장 = 1발주단위). 총수량은 수정배수가 곱해진 값이라 라벨 수량이 아님
            "qty":         ["발주단위수량", "총수량", "수량"],
            "date":        ["센터납품 예정일자", "예정일자", "고객인도일"],
        },
    },
    "gs": {
        "name": "GS",
        "tag": "GS",
        "title_tpl": "{jm}\n(GS슈퍼 GS스마트오더)",
        "center_label": "GS센터",
        "hub_col": "배송센터",
        "label_product": "raw",
        "col_width": {"A": 39.0, "B": 99.9, "C": 41.7, "D": 8.7},
        "row_h": 62.4,
        "lead_spacer": True,
        "raw": {
            "sheet_hint": ["lowdata", "발주", "Wine", "Sheet"],
            "header_scan_rows": 2,
            "center_code": ["센터코드"],
            "center_name": ["센터"],          # exact 우선 처리
            "store_code":  ["점포코드", "최초점포코드"],
            "store_name":  ["점포명"],
            "prod_code":   ["상품코드"],
            "prod_name":   ["상품명"],
            "qty":         ["수량"],
            "date":        ["예약일자", "수령일자"],
        },
    },
    "e24": {
        "name": "이마트24",
        "tag": "E-24",
        "title_tpl": "오비맥주 {jm}\n(E-24 스마트오더)",
        "center_label": "E-24센터",
        "hub_col": "한익스",         # E24 양식은 이고센터 컬럼명이 '한익스'
        "label_product": "raw",     # 라벨 상품명 = 화주상품명(원본). 표지는 대표(한익스상품명)
        "col_width": {"A": 47.8, "B": 110.7, "C": 41.7, "D": 8.7},
        "row_h": 83.4,
        "lead_spacer": False,
        "raw": {
            "sheet_hint": ["lowdata", "Sheet"],
            "header_scan_rows": 2,
            "center_code": ["입고센터명"],
            "center_name": ["입고센터명"],
            "store_code":  ["점포코드"],
            "store_name":  ["점포명"],
            "prod_code":   ["상품코드"],
            "prod_name":   ["앱 상품명", "상품명(대표)", "앱상품명", "상품명"],
            "qty":         ["주문수량", "수량"],
            "date":        ["납품예정일", "픽업일자", "발주일자"],
        },
    },
}

# 상품 구분(직매장 종류)
PRODUCT_CATS = ["대월", "프리미엄"]
DEFAULT_CAT = "대월"
TOTAL_TITLE_TPL = "{jm}({tag})스마트오더 TOTAL 수량"


def product_rep_cat(pm, code):
    """상품마스터에서 (대표=한익스상품명, 구분) 반환. 값이 문자열이면 레거시(구분=대월).
    미등록이면 (None, None)."""
    v = pm.get(code)
    if v is None:
        return None, None
    if isinstance(v, dict):
        return v.get("rep", ""), v.get("cat", DEFAULT_CAT)
    return v, DEFAULT_CAT


def make_title(cfg, cat):
    """상품 구분에 따른 라벨 제목. 대월→대월직매장, 프리미엄→프리미엄직매장."""
    jm = (cat or DEFAULT_CAT) + "직매장"
    return cfg["title_tpl"].format(jm=jm)


def cover_title(cfg, rows):
    """표지 제목. 포함된 구분이 하나면 그 구분, 여러 개면 '대월·프리미엄'."""
    cats = [r.get("구분") or DEFAULT_CAT for r in rows]
    uniq = sorted(set(cats), key=lambda c: (PRODUCT_CATS.index(c) if c in PRODUCT_CATS else 9))
    jm = ("·".join(uniq) if uniq else DEFAULT_CAT) + "직매장"
    return TOTAL_TITLE_TPL.format(jm=jm, tag=cfg["tag"])


# ----------------------------------------------------------------------------
# 유틸
# ----------------------------------------------------------------------------
def norm(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s

def to_int(v):
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except Exception:
        return v

# ----------------------------------------------------------------------------
# 마스터 로드/저장
# ----------------------------------------------------------------------------
def load_master(fname):
    p = os.path.join(MASTER_DIR, fname)
    if not os.path.exists(p):
        return {}
    with open(p, "r", encoding="utf-8") as f:
        return json.load(f)

def save_master(fname, data):
    p = os.path.join(MASTER_DIR, fname)
    with open(p, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1, sort_keys=True)

def all_hubs():
    """거점센터(이고센터) 후보 목록 — 관리 UI 드롭다운용."""
    return sorted({h for h in load_master(CENTER_MASTER).values() if h})


def all_reps():
    """등록된 한익스상품명 → 구분. 같은 상품이라도 채널마다 코드가 달라, 코드를 기존 상품에 붙일 때 쓴다."""
    pm = load_master("product.json")
    reps = {}
    for code in pm:
        rep, cat = product_rep_cat(pm, code)
        if rep:
            reps.setdefault(rep, cat or DEFAULT_CAT)
    return dict(sorted(reps.items()))

# ----------------------------------------------------------------------------
# 마스터 ↔ 엑셀 (다운로드 · 수정 · 업로드)
# ----------------------------------------------------------------------------
MASTER_KINDS = {
    "center": {
        "file": CENTER_MASTER,
        "headers": ["센터명", "이고센터"],
        "help": "센터명은 RAW에 찍히는 이름 그대로. 이고센터는 드롭다운에서 선택하세요.",
    },
    "product": {
        "file": "product.json",
        "headers": ["한익스상품명", "구분", "상품코드"],
        "help": "상품 1개 = 1행. 같은 상품이라도 채널마다 상품코드가 다르므로, "
                "코드가 여러 개면 한 칸에 쉼표(,)로 구분해 적으세요.",
    },
}

CODE_SEP = ", "


def split_codes(s):
    """'2202..., 2800...' → ['2202...', '2800...'] (쉼표/줄바꿈/공백 모두 허용)."""
    out = []
    for tok in norm(s).replace("\n", ",").replace(" ", ",").split(","):
        tok = tok.strip()
        if tok and tok not in out:
            out.append(tok)
    return out


def product_groups():
    """상품마스터를 상품 1개 = 1행으로. [{rep, cat, codes[]}] (한익스상품명 정렬)."""
    pm = load_master("product.json")
    groups = {}
    for code in sorted(pm):
        rep, cat = product_rep_cat(pm, code)
        g = groups.setdefault(rep or "", {"rep": rep or "", "cat": cat or DEFAULT_CAT, "codes": []})
        g["codes"].append(code)
    return sorted(groups.values(), key=lambda g: g["rep"])


def master_rows(kind):
    """마스터 → 엑셀 행 목록. 둘 다 '1행 = 1항목'."""
    if kind == "product":
        return [[g["rep"], g["cat"], CODE_SEP.join(g["codes"])] for g in product_groups()]
    data = load_master(MASTER_KINDS[kind]["file"])
    return sorted(([k, v] for k, v in data.items()), key=lambda r: (r[1], r[0]))


def build_master_xlsx(kind):
    """마스터 엑셀 다운로드. 이고센터/구분은 엑셀 드롭다운(목록 검증)으로 넣어 오타를 막는다."""
    from openpyxl.worksheet.datavalidation import DataValidation
    spec = MASTER_KINDS[kind]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = {"center": "센터마스터", "product": "상품마스터"}[kind]
    ws.append(spec["headers"])
    for c in range(1, len(spec["headers"]) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(name=FONT, size=11, bold=True)
        cell.fill = HDR_FILL
        cell.alignment = AL_C
    rows = master_rows(kind)
    for row in rows:
        ws.append(row)
    last = max(ws.max_row, 2) + 200   # 아래로 추가 입력할 여유 행까지 드롭다운 적용
    opts = all_hubs() if kind == "center" else PRODUCT_CATS
    col = "B"   # center: 이고센터, product: 구분 — 둘 다 두 번째 열
    if opts:
        dv = DataValidation(type="list", formula1='"%s"' % ",".join(opts), allow_blank=True)
        dv.error = "목록에 있는 값만 입력할 수 있습니다."
        ws.add_data_validation(dv)
        dv.add("%s2:%s%d" % (col, col, last))
    widths = {"center": [34, 16], "product": [34, 12, 46]}[kind]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(spec["headers"]) + 1):
            ws.cell(r, c).font = Font(name=FONT, size=11)
    ws.freeze_panes = "A2"
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def parse_master_xlsx(path, kind):
    """업로드된 마스터 엑셀 → 마스터 dict. (오류 행은 사유와 함께 함께 반환)"""
    spec = MASTER_KINDS[kind]
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    if not grid:
        raise ValueError("빈 파일입니다.")
    hdr = [norm(x) for x in grid[0]]
    need = spec["headers"]
    if hdr[:len(need)] != need:
        raise ValueError("첫 행 머리글이 달라요. '%s' 여야 합니다. (받은 값: %s)"
                         % (" / ".join(need), " / ".join(h for h in hdr[:len(need)] if h) or "빈 행"))
    data, errors = {}, []
    for i, r in enumerate(grid[1:], start=2):
        vals = [norm(r[c]) if c < len(r) else "" for c in range(len(need))]
        if not any(vals):
            continue
        if kind == "center":
            name, hub = vals
            if not name or not hub:
                errors.append("%d행: 센터명과 이고센터를 모두 채워주세요." % i); continue
            data[name] = hub
        else:
            rep, cat, codes_txt = vals
            codes = split_codes(codes_txt)
            if not rep or not codes:
                errors.append("%d행: 한익스상품명과 상품코드를 모두 채워주세요." % i); continue
            if cat and cat not in PRODUCT_CATS:
                errors.append("%d행: 구분은 %s 중 하나여야 합니다. (받은 값: %s)"
                              % (i, "/".join(PRODUCT_CATS), cat)); continue
            dup = [c for c in codes if c in data]
            if dup:
                errors.append("%d행: 상품코드 %s 가 다른 행에도 있습니다. 코드는 상품 하나에만 넣어주세요."
                              % (i, ", ".join(dup))); continue
            for code in codes:
                data[code] = {"rep": rep, "cat": cat or DEFAULT_CAT}
    return data, errors


def _master_val(kind, data, key):
    if kind == "product":
        rep, cat = product_rep_cat(data, key)
        return "%s (%s)" % (rep or "", cat or DEFAULT_CAT)
    return data.get(key, "")


def diff_master(kind, new):
    """현재 마스터 vs 업로드본 → 추가/수정/삭제 목록(적용 전 확인용)."""
    cur = load_master(MASTER_KINDS[kind]["file"])
    added, changed, removed = [], [], []
    for k in new:
        if k not in cur:
            added.append((k, _master_val(kind, new, k)))
        elif _master_val(kind, cur, k) != _master_val(kind, new, k):
            changed.append((k, _master_val(kind, cur, k), _master_val(kind, new, k)))
    for k in cur:
        if k not in new:
            removed.append((k, _master_val(kind, cur, k)))
    return {"added": sorted(added), "changed": sorted(changed), "removed": sorted(removed)}


# ----------------------------------------------------------------------------
# RAW 파싱
# ----------------------------------------------------------------------------
def _read_any(path):
    """xlsx/xls 모두 [ [row], ... ] 2차원 리스트로."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        return ws.name, [[ws.cell_value(i, j) for j in range(ws.ncols)] for i in range(ws.nrows)]
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    ws.reset_dimensions()   # 시트 크기를 A1:A1로 잘못 써두는 RAW가 있어(E24) 선언값을 믿지 않는다
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    name = ws.title
    wb.close()
    return name, rows

def _pick_sheet(path, hints):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(path)
        best = wb.sheet_by_index(0)
        for ws in wb.sheets():
            if any(h in ws.name for h in hints):
                best = ws; break
        def cval(i, j):
            # xls는 날짜가 실수(시리얼)로 저장됨 → datetime 복원(라벨출력 시트에 원본 날짜 유지)
            if best.cell_type(i, j) == xlrd.XL_CELL_DATE:
                try:
                    return xlrd.xldate.xldate_as_datetime(best.cell_value(i, j), wb.datemode)
                except Exception:
                    pass
            return best.cell_value(i, j)
        return best.name, [[cval(i, j) for j in range(best.ncols)] for i in range(best.nrows)]
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    target = wb.worksheets[0]
    for ws in wb.worksheets:
        if any(h in ws.title for h in hints):
            target = ws; break
    target.reset_dimensions()   # 시트 크기를 A1:A1로 잘못 써두는 RAW가 있어(E24) 선언값을 믿지 않는다
    rows = [list(r) for r in target.iter_rows(values_only=True)]
    name = target.title
    wb.close()
    return name, rows

def _find_header(rows, cfg):
    """헤더행 index와 각 필드 컬럼 index를 찾는다."""
    scan = cfg["raw"].get("header_scan_rows", 3)
    want = {k: cfg["raw"][k] for k in ("center_code","center_name","store_code","store_name","prod_code","prod_name","qty","date")}
    best_row, best_idx, best_hits = 0, {}, -1
    for r in range(min(scan + 1, len(rows))):
        hdr = [norm(x) for x in rows[r]]
        idx = {}
        for key, cands in want.items():
            # 후보를 '적힌 순서(우선순위)'대로 찾는다 — 컬럼 순서에 끌려가면 안 됨.
            # (예: CU는 총수량이 발주단위수량보다 앞 컬럼이지만, 수량은 발주단위수량이 정답)
            found = None
            for cand in cands:                       # 1) exact
                for i, h in enumerate(hdr):
                    if h == cand:
                        found = i; break
                if found is not None:
                    break
            if found is None:
                for cand in cands:                   # 2) 부분일치
                    for i, h in enumerate(hdr):
                        if cand in h:
                            found = i; break
                    if found is not None:
                        break
            if found is not None:
                idx[key] = found
        hits = len(idx)
        if hits > best_hits:
            best_row, best_idx, best_hits = r, idx, hits
    return best_row, best_idx

def parse_raw(path, channel):
    cfg = CHANNELS[channel]
    _, rows = _pick_sheet(path, cfg["raw"]["sheet_hint"])
    hrow, idx = _find_header(rows, cfg)
    required = ["center_name", "prod_code", "qty"]
    missing = [k for k in required if k not in idx]
    if missing:
        raise ValueError("RAW에서 필수 컬럼을 찾지 못했습니다: %s (헤더행 인식 실패)" % ", ".join(missing))

    def g(r, key):
        i = idx.get(key)
        if i is None or i >= len(r):
            return ""
        return norm(r[i])

    records = []
    for r in rows[hrow + 1:]:
        if not any(norm(x) for x in r):
            continue
        cname = g(r, "center_name")
        ccode = g(r, "center_code")
        pcode = g(r, "prod_code")
        pname = g(r, "prod_name")
        qtxt  = g(r, "qty")
        if not (cname or ccode) and not pcode:
            continue
        if not pcode and not pname:
            continue
        qty = to_int(qtxt) if qtxt else ""
        records.append({
            "center_code": ccode,
            "center_name": cname,
            "store_code": g(r, "store_code"),
            "store_name": g(r, "store_name") or cname,
            "prod_code": pcode,
            "prod_name": pname,
            "qty": qty,
            "date": g(r, "date"),
        })
    return records

# ----------------------------------------------------------------------------
# 매핑 처리
# ----------------------------------------------------------------------------
def process(records, channel):
    """레코드에 거점센터/대표상품명을 코드 기반으로 매핑.
    반환: (rows, unmapped_centers, unmapped_products)
    rows: 붙여넣기(데이터추출)용 dict 리스트
    """
    cfg = CHANNELS[channel]
    center_master = load_master(CENTER_MASTER)
    product_master = load_master("product.json")
    # 보조: 화주상품명 -> {rep,cat} (코드 미등록 시 이름 폴백). 정규화 키로 인덱싱.
    product_names = load_master("product_names.json")
    pn_idx = {norm(k): v for k, v in product_names.items()}

    rows = []
    unmapped_centers = {}   # 센터명 -> 참고표시(센터코드)
    unmapped_products = {}  # code -> name
    for rec in records:
        ckeyval = rec["center_name"]   # 센터 매핑 키 = 센터명(전 채널 공통)
        hub = center_master.get(ckeyval, "")
        if not hub and ckeyval:
            unmapped_centers[ckeyval] = rec["center_code"]
        # 1) 코드 우선 → 2) 화주상품명 폴백
        rep, cat = product_rep_cat(product_master, rec["prod_code"])
        if rep is None:
            v = pn_idx.get(norm(rec["prod_name"]))
            if v:
                rep, cat = v.get("rep", ""), v.get("cat", DEFAULT_CAT)
        if not rep and rec["prod_code"]:
            unmapped_products[rec["prod_code"]] = rec["prod_name"]
        cat = cat or DEFAULT_CAT
        # 부착양식 라벨 = 화주상품명(원본 제품명), 표지/집계 = 대표(한익스상품명)
        label_prod = rec["prod_name"] if cfg["label_product"] == "raw" else (rep or rec["prod_name"])
        rows.append({
            "센터": rec["center_name"],
            "배송일자": rec["date"],
            "점포코드": rec["store_code"],
            "점포명": rec["store_name"],
            "거점센터": hub,
            "상품명": label_prod,          # 라벨 표시용 = 화주상품명
            "수량": rec["qty"],
            # 표지 표시용 = 한익스상품명(대표). 마스터에 없으면 빈값 — 화주상품명으로 폴백하지 않는다.
            "대표": rep or "",
            "구분": cat,                       # 대월 / 프리미엄 (라벨 제목 분기)
            "_ckey": ckeyval,
            "_pcode": rec["prod_code"],
        })
    return rows, unmapped_centers, unmapped_products


def missing_reps(rows):
    """한익스상품명(대표)이 비어 있는 행 → {상품코드: 화주상품명}. 비어 있지 않으면 출력 차단 대상."""
    miss = {}
    for row in rows:
        if not row["대표"]:
            miss[row["_pcode"] or row["상품명"]] = row["상품명"]
    return miss


def sort_rows(rows):
    """구분(대월/프리미엄) → 상품명(화주·라벨표시) → 이고센터 → 채널센터 → 점포명 순 정렬."""
    def catkey(c):
        return PRODUCT_CATS.index(c) if c in PRODUCT_CATS else 9
    return sorted(rows, key=lambda x: (catkey(x.get("구분") or DEFAULT_CAT),
                                       x["상품명"] or "",
                                       x["거점센터"] or "zzz",
                                       x["센터"] or "",
                                       x["점포명"] or ""))


REP_COL = "상품명(대표)"   # 라벨출력(폼텍) 시트의 한익스상품명 컬럼명


def labelout_grid(path, channel):
    """폼텍 디자인프로용 '라벨출력' 그리드.

    RAW 원본 컬럼을 그대로 보존하고 뒤에 매핑 컬럼 2개를 덧붙인다.
      - {hub_col}: 이고센터 (CU/GS='배송센터', E24='한익스')
      - 상품명(대표): 한익스상품명
    행 순서는 부착양식과 동일(구분 → 상품명 → 이고센터 → 채널센터).
    반환: (headers, rows) — rows의 셀은 RAW 원본 값(날짜는 datetime 유지).
    """
    cfg = CHANNELS[channel]
    _, grid = _pick_sheet(path, cfg["raw"]["sheet_hint"])
    hrow, idx = _find_header(grid, cfg)
    if "prod_code" not in idx or "qty" not in idx:
        raise ValueError("RAW에서 필수 컬럼을 찾지 못했습니다(라벨출력).")

    headers = [h for h in grid[hrow]]
    while headers and not norm(headers[-1]):   # 서식만 있는 꼬리 빈 컬럼 제거(E24 양식)
        headers.pop()
    ncol = len(headers)
    hub_col, rep_col = cfg["hub_col"], REP_COL
    # 이미 같은 컬럼이 있으면(출력양식 재업로드) 덧붙이지 않고 덮어쓴다
    hdr_norm = [norm(h) for h in headers]
    hub_i = hdr_norm.index(hub_col) if hub_col in hdr_norm else None
    rep_i = hdr_norm.index(rep_col) if rep_col in hdr_norm else None
    if hub_i is None:
        headers.append(hub_col); hub_i = ncol; ncol += 1
    if rep_i is None:
        headers.append(rep_col); rep_i = ncol; ncol += 1

    center_master = load_master(CENTER_MASTER)
    product_master = load_master("product.json")
    pn_idx = {norm(k): v for k, v in load_master("product_names.json").items()}

    def g(r, key):
        i = idx.get(key)
        if i is None or i >= len(r):
            return ""
        return norm(r[i])

    out = []
    for r in grid[hrow + 1:]:
        if not any(norm(x) for x in r):
            continue
        cname, ccode = g(r, "center_name"), g(r, "center_code")
        pcode, pname = g(r, "prod_code"), g(r, "prod_name")
        if not (cname or ccode) and not pcode:
            continue
        if not pcode and not pname:
            continue
        hub = center_master.get(cname, "")
        rep, cat = product_rep_cat(product_master, pcode)
        if not rep:
            v = pn_idx.get(norm(pname))
            if v:
                rep, cat = v.get("rep", ""), v.get("cat", DEFAULT_CAT)
        cells = list(r) + [None] * (ncol - len(r))
        cells[hub_i] = hub
        cells[rep_i] = rep or ""
        out.append({"cells": cells[:ncol], "_cat": cat or DEFAULT_CAT,
                    "_prod": pname, "_hub": hub, "_center": cname,
                    "_store": g(r, "store_name") or cname})

    def catkey(c):
        return PRODUCT_CATS.index(c) if c in PRODUCT_CATS else 9
    out.sort(key=lambda x: (catkey(x["_cat"]), x["_prod"] or "",
                            x["_hub"] or "zzz", x["_center"] or "", x["_store"] or ""))
    return headers, [x["cells"] for x in out]


HDR_FILL = PatternFill("solid", fgColor="FFEEF2F5")


def build_labelout_sheet(ws, headers, rows):
    """폼텍 디자인프로 데이터 원본 시트(1행=헤더, 1행=라벨 1장)."""
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(name=FONT, size=10)
        cell.fill = HDR_FILL
        cell.alignment = AL_C
    for row in rows:
        ws.append(row)
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.font = Font(name=FONT, size=10)
            if isinstance(cell.value, (datetime.datetime, datetime.date)):
                cell.number_format = "yyyy-mm-dd"
    for c, h in enumerate(headers, 1):
        w = max(9, min(28, len(norm(h)) + 6))
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"


def compute_totals(rows):
    """TOTAL 표지용 집계: 상품(대표)별 수량, 거점센터별 수량. (이름 오름차순)"""
    prod, hub = {}, {}
    for row in rows:
        q = row["수량"] if isinstance(row["수량"], (int, float)) else 0
        if row["대표"]:
            prod[row["대표"]] = prod.get(row["대표"], 0) + q
        if row["거점센터"]:
            hub[row["거점센터"]] = hub.get(row["거점센터"], 0) + q
    prod_list = sorted(prod.items())
    hub_list = sorted(hub.items())
    return prod_list, hub_list

# ----------------------------------------------------------------------------
# 스타일 (기존 부착양식과 동일)
# ----------------------------------------------------------------------------
FONT = "맑은 고딕"
_thin = Side(style="thin", color="FF000000")
BORDER_BOX = Border(top=_thin, bottom=_thin, left=_thin, right=_thin)
FILL_YELLOW = PatternFill("solid", fgColor="FFFFFF00")
FILL_WARN = PatternFill("solid", fgColor=Color(theme=1, tint=0.05))
AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_C_NW = Alignment(horizontal="center", vertical="center", wrap_text=False)
# 값 셀: 계산한 글자크기 + 엑셀 '축소하여 전체 표시'(추정이 빗나가도 잘리지 않게 하는 안전망)
AL_C_FIT = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=True)
AL_V = Alignment(vertical="center")
AL_CNUM = Alignment(vertical="center", wrap_text=True)   # 노란 C열(라벨번호) — 원본과 동일

def _text_units(s):
    """반각 1 / 한글·한자 등 전각 2 로 센 텍스트 폭(반각 단위)."""
    n = 0
    for ch in norm(s):
        n += 2 if ord(ch) > 0x1100 else 1
    return n


def _col_px(width):
    """엑셀 열 너비(문자 단위) → 픽셀."""
    return width * 7 + 5


def fit_font_size(text, col_width, base, min_size=22, pad_px=12):
    """열 너비 안에 한 줄로 다 들어가는 최대 글자크기(pt). base를 넘지 않고 min_size 아래로도 안 내려감.
    전각 1글자 폭 ≈ 글자크기(pt) × 96/72 픽셀 로 추정."""
    units = _text_units(text)
    if units == 0:
        return base
    avail = _col_px(col_width) - pad_px
    size = avail / (units * (96 / 72) / 2)
    return max(min_size, min(base, int(size)))


def _cnum_cell(cell, value=None):
    cell.value = value
    cell.font = Font(name=FONT, size=48, bold=True)
    cell.fill = FILL_YELLOW
    cell.alignment = AL_CNUM

def _style_cell(cell, value, size, bold=True, border=True, fill=None, align=AL_C_NW):
    cell.value = value
    cell.font = Font(name=FONT, size=size, bold=bold)
    cell.alignment = align
    if border:
        cell.border = BORDER_BOX
    if fill:
        cell.fill = fill

def build_label_sheet(ws, rows, cfg):
    # 열 너비
    for col, w in cfg["col_width"].items():
        ws.column_dimensions[col].width = w
    r = 1
    if cfg["lead_spacer"]:
        ws.row_dimensions[r].height = 25.5
        _cnum_cell(ws.cell(r, 3))
        r += 1
    rh = cfg["row_h"]
    wa, wbv = cfg["col_width"]["A"], cfg["col_width"]["B"]
    wtitle = wa + wbv          # 제목은 A:B 병합
    for n, row in enumerate(rows, start=1):
        # 1) 제목 (상품 구분에 따라 대월/프리미엄 직매장)
        title = make_title(cfg, row.get("구분"))
        tsize = min(fit_font_size(line, wtitle, 62, min_size=40) for line in title.split("\n"))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        _style_cell(ws.cell(r, 1), title, tsize, fill=None, align=AL_C)
        _cnum_cell(ws.cell(r, 3), n)
        ws.row_dimensions[r].height = 169.95
        r += 1
        # 2) 개봉금지
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        _style_cell(ws.cell(r, 1), "※수령자 외 절대 개봉금지※", 40, fill=FILL_WARN, align=AL_C)
        _cnum_cell(ws.cell(r, 3))
        ws.row_dimensions[r].height = 63.6
        r += 1
        # 3~7) 라벨 항목
        items = [
            ("이고센터", row["거점센터"], 42),
            (cfg["center_label"], row["센터"], 42),
            (" 점포명", row["점포명"], 48),
            (" 상품 ", row["상품명"], 48),
            (" 수량", row["수량"], 48),
        ]
        for label, value, bsize in items:
            # 긴 상품명·센터명이 열 밖으로 잘리지 않도록 글자크기를 폭에 맞춰 축소
            _style_cell(ws.cell(r, 1), label, fit_font_size(label, wa, 42), align=AL_C_FIT)
            _style_cell(ws.cell(r, 2), value, fit_font_size(value, wbv, bsize), align=AL_C_FIT)
            _cnum_cell(ws.cell(r, 3))
            ws.row_dimensions[r].height = rh
            r += 1
        # 8) 여백
        cc = ws.cell(r, 3); cc.font = Font(name=FONT, size=48, bold=True); cc.fill = FILL_YELLOW; cc.alignment = AL_V
        ws.row_dimensions[r].height = 26.0
        r += 1

def build_data_sheet(ws, rows):
    headers = ["센터", "배송일자", "점포코드", "점포명", "거점센터", "상품명", "수량", "상품명(대표)"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = Font(name=FONT, size=11, bold=True)
    for row in rows:
        ws.append([row["센터"], row["배송일자"], row["점포코드"], row["점포명"],
                   row["거점센터"], row["상품명"], row["수량"], row["대표"]])
    widths = [22, 12, 10, 24, 14, 30, 8, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def build_total_sheet(ws, rows, cfg):
    from collections import OrderedDict
    prod = OrderedDict()
    hub = OrderedDict()
    for row in rows:
        q = row["수량"] if isinstance(row["수량"], (int, float)) else 0
        if row["대표"]:   # 표지 상품명은 항상 한익스상품명(대표)만
            prod[row["대표"]] = prod.get(row["대표"], 0) + q
        if row["거점센터"]:
            hub[row["거점센터"]] = hub.get(row["거점센터"], 0) + q
    ws.cell(1, 2, cover_title(cfg, rows)).font = Font(name=FONT, size=14, bold=True)
    hdr = ["번호", "상품", "수량"]
    for i, h in enumerate(hdr):
        ws.cell(2, 2 + i, h).font = Font(name=FONT, size=11, bold=True)
    ws.cell(2, 6, "이고센터명").font = Font(name=FONT, size=11, bold=True)
    ws.cell(2, 7, "이고수량").font = Font(name=FONT, size=11, bold=True)
    r = 3
    for i, (k, v) in enumerate(sorted(prod.items()), 1):   # 이름 오름차순
        ws.cell(r, 2, i); ws.cell(r, 3, k); ws.cell(r, 4, v); r += 1
    r = 3
    for k, v in sorted(hub.items()):
        ws.cell(r, 6, k); ws.cell(r, 7, v); r += 1
    for col, w in {"B": 6, "C": 34, "D": 8, "E": 3, "F": 16, "G": 10}.items():
        ws.column_dimensions[col].width = w

def generate_labelout_workbook(path, channel):
    """폼텍 디자인프로 전용: 'lowdata(라벨출력)' 시트 하나만 담은 xlsx."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "lowdata(라벨출력)"
    headers, grid = labelout_grid(path, channel)
    build_labelout_sheet(ws, headers, grid)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def generate_workbook(rows, channel, path=None):
    """path를 주면 폼텍용 'lowdata(라벨출력)' 시트를 첫 탭으로 함께 생성(RAW 원본 + 매핑 컬럼)."""
    cfg = CHANNELS[channel]
    rows = sort_rows(rows)
    wb = openpyxl.Workbook()
    try:
        wb._named_styles["Normal"].font = Font(name=FONT, size=11)  # 빈 셀 기본폰트도 맑은 고딕
    except Exception:
        pass
    ws_first = wb.active
    if path:
        ws_first.title = "lowdata(라벨출력)"
        headers, grid = labelout_grid(path, channel)
        build_labelout_sheet(ws_first, headers, grid)
        ws_total = wb.create_sheet("TOTAL(표지)")
    else:
        ws_total = ws_first
        ws_total.title = "TOTAL(표지)"
    build_total_sheet(ws_total, rows, cfg)
    ws_label = wb.create_sheet("부착양식(A4출력)")
    build_label_sheet(ws_label, rows, cfg)
    ws_data = wb.create_sheet("붙여넣기(데이터추출)")
    build_data_sheet(ws_data, rows)
    # 인쇄 설정: A4 세로, 여백 최소, 열 폭 맞춤
    ws_label.page_setup.orientation = "portrait"
    ws_label.page_setup.paperSize = 9  # A4
    ws_label.page_setup.fitToWidth = 1
    ws_label.page_setup.fitToHeight = 0
    ws_label.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws_label.print_area = "A1:C%d" % ws_label.max_row
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
