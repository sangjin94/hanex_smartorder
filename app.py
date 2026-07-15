# -*- coding: utf-8 -*-
"""한익스프레스 스마트오더 라벨 생성기 (웹)

기능
 - 판매채널(CU/GS/E24)별 RAW 업로드 → 코드 기반 매핑 → 부착양식(A4) 라벨 xlsx 생성
 - 업로드 중 미매핑 센터/상품 코드를 화면에서 즉시 마스터에 등록
 - 마스터 관리(센터명→이고센터[전 채널 공통], 상품코드→한익스상품명) 조회/추가/수정/삭제
"""
import os, io, json, re, uuid, datetime, sys

# --- 단일 실행파일(.exe, PyInstaller) 지원 ---------------------------------
# 리소스(templates/static/masters seed)는 번들에서 읽고,
# 사용자 데이터(마스터 수정본·업로드 아카이브)는 exe 옆 data/ 폴더에 저장(영구·폴더복사 가능).
if getattr(sys, "frozen", False):
    _BUNDLE = sys._MEIPASS
    _APPDIR = os.path.dirname(sys.executable)
    os.environ.setdefault("SMARTORDER_SEED", os.path.join(_BUNDLE, "masters"))
    os.environ.setdefault("SMARTORDER_MASTERS", os.path.join(_APPDIR, "data", "masters"))
    os.environ.setdefault("SMARTORDER_ARCHIVE", os.path.join(_APPDIR, "data", "archive"))
    _TPL = os.path.join(_BUNDLE, "templates")
    _STATIC = os.path.join(_BUNDLE, "static")
else:
    _TPL, _STATIC = "templates", "static"

from flask import (Flask, request, render_template, send_file, redirect,
                   url_for, flash, jsonify, abort)
import smartorder_core as sc

app = Flask(__name__, template_folder=_TPL, static_folder=_STATIC)
app.secret_key = "hanex-smartorder-label-2026"
app.config["MAX_CONTENT_LENGTH"] = 40 * 1024 * 1024  # 40MB

# nginx 서브경로(/smartorder/) 배포 시 url_for가 접두어를 붙이도록 처리
class PrefixMiddleware:
    def __init__(self, wsgi_app):
        self.app = wsgi_app
    def __call__(self, environ, start_response):
        prefix = environ.get("HTTP_X_FORWARDED_PREFIX", "")
        if prefix:
            environ["SCRIPT_NAME"] = prefix.rstrip("/")
        return self.app(environ, start_response)
app.wsgi_app = PrefixMiddleware(app.wsgi_app)

sc.ensure_masters_seeded()

CH_ORDER = ["cu", "gs", "e24"]

# ---------------- 업로드 아카이브(영구 누적) ----------------
# 배포 시 git 밖 영구 디렉터리로 분리(SMARTORDER_ARCHIVE). 기본은 앱 폴더 archive/.
ARCHIVE_DIR = os.environ.get("SMARTORDER_ARCHIVE") or os.path.join(sc.BASE, "archive")
INDEX_PATH = os.path.join(ARCHIVE_DIR, "index.json")
os.makedirs(ARCHIVE_DIR, exist_ok=True)

def _load_index():
    if not os.path.exists(INDEX_PATH):
        return []
    try:
        with open(INDEX_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []

def _save_index(items):
    with open(INDEX_PATH, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=1)

def _safe_name(name):
    base = os.path.basename(name)
    return re.sub(r'[\\/:*?"<>|]+', "_", base).strip() or "raw"

def _rec_by_id(uid):
    for r in _load_index():
        if r["id"] == uid:
            return r
    return None

def _rec_path(rec):
    return os.path.join(ARCHIVE_DIR, rec["ch"], rec["fname"])


@app.route("/")
def index():
    centers = len(sc.load_master(sc.CENTER_MASTER))
    products = len(sc.load_master("product.json"))
    return render_template("index.html", channels=CH_ORDER, cfg=sc.CHANNELS,
                           centers=centers, products=products)


@app.route("/u/<ch>", methods=["GET", "POST"])
def upload(ch):
    if ch not in sc.CHANNELS:
        abort(404)
    cfg = sc.CHANNELS[ch]
    if request.method == "GET":
        recent = [r for r in reversed(_load_index()) if r["ch"] == ch][:8]
        return render_template("upload.html", ch=ch, cfg=cfg, recent=recent)

    f = request.files.get("file")
    if not f or not f.filename:
        flash("파일을 선택해 주세요.", "error")
        return redirect(url_for("upload", ch=ch))
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in (".xlsx", ".xls"):
        flash("xlsx 또는 xls 파일만 업로드할 수 있습니다.", "error")
        return redirect(url_for("upload", ch=ch))

    # 채널 폴더에 영구 저장(누적)
    now = datetime.datetime.now()
    uid = "%s_%s_%s" % (ch, now.strftime("%Y%m%d_%H%M%S"), uuid.uuid4().hex[:4])
    fname = uid + "__" + _safe_name(f.filename)
    os.makedirs(os.path.join(ARCHIVE_DIR, ch), exist_ok=True)
    path = os.path.join(ARCHIVE_DIR, ch, fname)
    f.save(path)

    # 통계 미리 계산(이력 표시용) — 실패해도 파일은 보관
    nrec = nrows = total_qty = None
    try:
        records = sc.parse_raw(path, ch)
        rows, uc, up = sc.process(records, ch)
        nrec, nrows = len(records), len(rows)
        total_qty = sum(r["수량"] for r in rows if isinstance(r["수량"], (int, float)))
    except Exception:
        pass

    idx = _load_index()
    idx.append({"id": uid, "ch": ch, "orig": f.filename, "fname": fname,
                "uploaded_at": now.strftime("%Y-%m-%d %H:%M:%S"),
                "nrec": nrec, "nrows": nrows, "total_qty": total_qty})
    _save_index(idx)
    return redirect(url_for("result", ch=ch, token=uid))


def _load_job(ch, token):
    rec = _rec_by_id(token)
    if not rec or rec["ch"] != ch:
        return None
    path = _rec_path(rec)
    if not os.path.exists(path):
        return None
    return {"channel": ch, "path": path, "filename": rec["orig"]}


@app.route("/r/<ch>/<token>")
def result(ch, token):
    if ch not in sc.CHANNELS:
        abort(404)
    job = _load_job(ch, token)
    if not job:
        flash("업로드 세션이 만료되었습니다. 다시 업로드해 주세요.", "error")
        return redirect(url_for("upload", ch=ch))
    cfg = sc.CHANNELS[ch]
    try:
        records = sc.parse_raw(job["path"], ch)
    except Exception as e:
        flash("RAW 파일 분석 실패: %s" % e, "error")
        return redirect(url_for("upload", ch=ch))
    rows, uc, up = sc.process(records, ch)
    rows = sc.sort_rows(rows)
    total_qty = sum(r["수량"] for r in rows if isinstance(r["수량"], (int, float)))
    prod_list, hub_list = sc.compute_totals(rows)
    hubs = sc.all_hubs()
    # 기존 한익스상품명 목록(+구분) — 미매핑 코드를 기존 상품에 붙일 때 오타로 새 상품이 생기는 걸 방지
    rep_cats = sc.all_reps()
    return render_template("result.html", ch=ch, cfg=cfg, token=token,
                           filename=job["filename"], nrec=len(records),
                           nrows=len(rows), total_qty=total_qty,
                           unmapped_centers=uc, unmapped_products=up,
                           hubs=hubs, rep_cats=rep_cats,
                           prod_list=prod_list, hub_list=hub_list,
                           cover_title=sc.cover_title(cfg, rows),
                           product_cats=sc.PRODUCT_CATS)


def _block_if_missing_reps(rows, ch, token):
    """한익스상품명(대표)이 없는 상품이 하나라도 있으면 출력(다운로드·인쇄)을 막는다.
    표지에는 항상 한익스상품명만 들어가야 하므로, 화주상품명으로 대체 출력하지 않는다."""
    miss = sc.missing_reps(rows)
    if not miss:
        return None
    flash("한익스상품명이 등록되지 않은 상품이 있어 출력할 수 없습니다: %s — 아래에서 등록 후 다시 시도해 주세요."
          % ", ".join("%s(%s)" % (c, n) for c, n in list(miss.items())[:5]), "error")
    return redirect(url_for("result", ch=ch, token=token))


@app.route("/print/<ch>/<token>")
def print_labels(ch, token):
    if ch not in sc.CHANNELS:
        abort(404)
    job = _load_job(ch, token)
    if not job:
        flash("업로드 세션이 만료되었습니다.", "error")
        return redirect(url_for("upload", ch=ch))
    cfg = sc.CHANNELS[ch]
    records = sc.parse_raw(job["path"], ch)
    rows, uc, up = sc.process(records, ch)
    blocked = _block_if_missing_reps(rows, ch, token)
    if blocked:
        return blocked
    rows = sc.sort_rows(rows)
    for r in rows:
        r["title"] = sc.make_title(cfg, r.get("구분"))
    return render_template("print.html", ch=ch, cfg=cfg, rows=rows, token=token)


@app.route("/assign/<ch>/<token>", methods=["POST"])
def assign(ch, token):
    if ch not in sc.CHANNELS:
        abort(404)
    # 센터 매핑 저장(전 채널 공통 마스터, 키=센터명)
    cm = sc.load_master(sc.CENTER_MASTER)
    for key in request.form:
        if key.startswith("center::"):
            name = key[len("center::"):]
            val = request.form.get(key, "").strip()
            if val:
                cm[name] = val
    sc.save_master(sc.CENTER_MASTER, cm)
    # 상품 매핑 저장 (공통 마스터) — {대표(한익스), 구분}
    pm = sc.load_master("product.json")
    for key in request.form:
        if key.startswith("prod::"):
            code = key[len("prod::"):]
            rep = request.form.get(key, "").strip()
            cat = request.form.get("pcat::" + code, sc.DEFAULT_CAT).strip() or sc.DEFAULT_CAT
            if rep:
                pm[code] = {"rep": rep, "cat": cat}
    sc.save_master("product.json", pm)
    flash("매핑을 저장했습니다.", "ok")
    return redirect(url_for("result", ch=ch, token=token))


@app.route("/download/<ch>/<token>")
def download(ch, token):
    if ch not in sc.CHANNELS:
        abort(404)
    job = _load_job(ch, token)
    if not job:
        flash("업로드 세션이 만료되었습니다.", "error")
        return redirect(url_for("upload", ch=ch))
    cfg = sc.CHANNELS[ch]
    records = sc.parse_raw(job["path"], ch)
    rows, uc, up = sc.process(records, ch)
    blocked = _block_if_missing_reps(rows, ch, token)
    if blocked:
        return blocked
    bio = sc.generate_workbook(rows, ch, job["path"])
    today = datetime.datetime.now().strftime("%Y%m%d")
    fname = "%s_스마트오더_부착양식_%s.xlsx" % (cfg["name"], today)
    return send_file(bio, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/labelout/<ch>/<token>")
def labelout(ch, token):
    """폼텍 디자인프로 데이터 원본용: RAW 원본 컬럼 + 배송센터(이고) + 상품명(대표) 시트만 담은 xlsx."""
    if ch not in sc.CHANNELS:
        abort(404)
    job = _load_job(ch, token)
    if not job:
        flash("업로드 세션이 만료되었습니다.", "error")
        return redirect(url_for("upload", ch=ch))
    cfg = sc.CHANNELS[ch]
    records = sc.parse_raw(job["path"], ch)
    rows, uc, up = sc.process(records, ch)
    blocked = _block_if_missing_reps(rows, ch, token)
    if blocked:
        return blocked
    bio = sc.generate_labelout_workbook(job["path"], ch)
    today = datetime.datetime.now().strftime("%Y%m%d")
    fname = "%s_스마트오더_라벨출력(폼텍)_%s.xlsx" % (cfg["name"], today)
    return send_file(bio, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------- 업로드 이력(누적 파일) ----------------
@app.route("/history")
def history():
    ch = request.args.get("ch", "")
    items = list(reversed(_load_index()))
    if ch in sc.CHANNELS:
        items = [r for r in items if r["ch"] == ch]
    counts = {}
    for r in _load_index():
        counts[r["ch"]] = counts.get(r["ch"], 0) + 1
    return render_template("history.html", items=items, cfg=sc.CHANNELS,
                           channels=CH_ORDER, sel=ch, counts=counts)


@app.route("/history/raw/<uid>")
def history_raw(uid):
    rec = _rec_by_id(uid)
    if not rec:
        abort(404)
    path = _rec_path(rec)
    if not os.path.exists(path):
        abort(404)
    return send_file(path, as_attachment=True, download_name=rec["orig"])


@app.route("/history/delete/<uid>", methods=["POST"])
def history_delete(uid):
    idx = _load_index()
    rec = next((r for r in idx if r["id"] == uid), None)
    if rec:
        try:
            os.remove(_rec_path(rec))
        except Exception:
            pass
        idx = [r for r in idx if r["id"] != uid]
        _save_index(idx)
        flash("삭제했습니다: %s" % rec["orig"], "ok")
    return redirect(url_for("history", ch=request.form.get("ch", "")))


# ---------------- 작업 집계(검수리스트): 채널·일자별 · 이고센터>센터>상품(화주)>수량 ----------------
def _norm_date(s):
    """'20260715' / '2026-07-14 00:00:00' / '2026-07-14' → '2026-07-15'. 없으면 ''"""
    s = (s or "").strip()
    if not s:
        return ""
    digits = re.sub(r"[^0-9]", "", s.split(" ")[0])
    if len(digits) >= 8:
        return "%s-%s-%s" % (digits[0:4], digits[4:6], digits[6:8])
    return s


def _hubkey(h):
    return (1, "") if h in ("(미지정)", "") else (0, h)


def _checklist(ch, sel_date):
    """한 채널의 업로드들을 (이고센터 > 채널센터 > 화주상품명) 별 수량으로 집계.
    수량은 process()의 row['수량'](CU=발주단위수량). 반환: (blocks, grand, dates)"""
    agg = {}          # (hub, center, prod) -> qty
    dates = set()
    for rec in _load_index():
        if rec["ch"] != ch:
            continue
        path = _rec_path(rec)
        if not os.path.exists(path):
            continue
        try:
            rows, _, _ = sc.process(sc.parse_raw(path, ch), ch)
        except Exception:
            continue
        for row in rows:
            d = _norm_date(row["배송일자"]) or "(일자없음)"
            dates.add(d)
            if sel_date and d != sel_date:
                continue
            q = row["수량"] if isinstance(row["수량"], (int, float)) else 0
            key = (row["거점센터"] or "(미지정)", row["센터"] or "", row["상품명"] or "")
            agg[key] = agg.get(key, 0) + q

    # 트리: hub -> center -> [(prod, qty)]
    from collections import OrderedDict
    tree = OrderedDict()
    for (hub, center, prod) in sorted(agg, key=lambda k: (_hubkey(k[0]), k[1], k[2])):
        tree.setdefault(hub, OrderedDict()).setdefault(center, []).append((prod, agg[(hub, center, prod)]))

    blocks, grand = [], 0
    for hub, centers in tree.items():
        cblocks, hub_total = [], 0
        for center, prods in centers.items():
            sub = sum(q for _, q in prods)
            hub_total += sub
            cblocks.append({"center": center, "prods": prods, "sub": sub})
        grand += hub_total
        blocks.append({"hub": hub, "centers": cblocks, "hub_total": hub_total})
    return blocks, grand, sorted(dates, reverse=True)


# --- 상품집계(이전 방식): 일자 · 한익스상품명 기준, 채널별/거점센터별 ---
def _aggregate():
    """아카이브 전체 → (일자, 한익스상품명, 구분)별 채널·거점센터 수량. skipped=한익스상품명 없는 수량."""
    agg, dates, hubs, skipped = {}, set(), set(), 0
    for rec in _load_index():
        ch, path = rec["ch"], _rec_path(rec)
        if not os.path.exists(path):
            continue
        try:
            rows, _, _ = sc.process(sc.parse_raw(path, ch), ch)
        except Exception:
            continue
        for row in rows:
            q = row["수량"] if isinstance(row["수량"], (int, float)) else 0
            rep = row["대표"]
            d = _norm_date(row["배송일자"]) or "(일자없음)"
            if not rep:
                skipped += q
                continue
            hub = row["거점센터"] or "(거점미지정)"
            dates.add(d); hubs.add(hub)
            e = agg.setdefault((d, rep, row["구분"]), {"ch": {}, "hub": {}})
            e["ch"][ch] = e["ch"].get(ch, 0) + q
            e["hub"][hub] = e["hub"].get(hub, 0) + q
    return agg, sorted(dates, reverse=True), sorted(hubs), skipped


def _report_table(view, sel_date):
    agg, dates, hubs, skipped = _aggregate()
    if view == "hub":
        columns = [(h, h) for h in hubs]
    else:
        columns = [(c, sc.CHANNELS[c]["name"]) for c in CH_ORDER]
    colkeys = [c for c, _ in columns]
    groups = {}
    for (d, rep, cat), e in agg.items():
        if sel_date and d != sel_date:
            continue
        src = e["hub"] if view == "hub" else e["ch"]
        groups.setdefault(d, []).append({"rep": rep, "cat": cat,
                                         "cols": {k: src.get(k, 0) for k in colkeys},
                                         "total": sum(src.values())})
    out, grand, grand_total = [], {k: 0 for k in colkeys}, 0
    for d in sorted(groups, reverse=True):
        rows = sorted(groups[d], key=lambda r: (sc.PRODUCT_CATS.index(r["cat"]) if r["cat"] in sc.PRODUCT_CATS else 9, r["rep"]))
        sub = {k: sum(r["cols"][k] for r in rows) for k in colkeys}
        sub_total = sum(r["total"] for r in rows)
        for k in colkeys:
            grand[k] += sub[k]
        grand_total += sub_total
        out.append({"date": d, "rows": rows, "sub": sub, "sub_total": sub_total})
    return columns, out, grand, grand_total, dates, skipped


@app.route("/report")
def report():
    mode = "product" if request.args.get("mode") == "product" else "checklist"
    sel_date = request.args.get("date", "")
    if mode == "product":
        view = "hub" if request.args.get("view") == "hub" else "channel"
        columns, groups, grand, grand_total, dates, skipped = _report_table(view, sel_date)
        return render_template("report.html", mode="product", view=view, sel_date=sel_date,
                               columns=columns, groups=groups, grand=grand, grand_total=grand_total,
                               dates=dates, skipped=skipped, cfg=sc.CHANNELS, channels=CH_ORDER, ch="")
    ch = request.args.get("ch", "")
    if ch not in sc.CHANNELS:
        ch = CH_ORDER[0]
    blocks, grand, dates = _checklist(ch, sel_date)
    return render_template("report.html", mode="checklist", ch=ch, cfg=sc.CHANNELS, channels=CH_ORDER,
                           sel_date=sel_date, blocks=blocks, grand=grand, dates=dates)


@app.route("/report/xlsx")
def report_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    mode = "product" if request.args.get("mode") == "product" else "checklist"
    sel_date = request.args.get("date", "")
    yellow = PatternFill("solid", fgColor="FFFFFF00")
    thin = Side(style="thin", color="FFBFBFBF")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    wb = openpyxl.Workbook(); ws = wb.active
    today = datetime.datetime.now().strftime("%Y%m%d")

    if mode == "product":
        view = "hub" if request.args.get("view") == "hub" else "channel"
        columns, groups, grand, grand_total, dates, skipped = _report_table(view, sel_date)
        ws.title = "상품집계"
        headers = ["일자", "한익스상품명", "구분"] + [name for _, name in columns] + ["합계"]
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            ws.cell(1, c).font = Font(name="맑은 고딕", size=11, bold=True); ws.cell(1, c).fill = yellow
        for g in groups:
            for r in g["rows"]:
                ws.append([g["date"], r["rep"], r["cat"]] + [r["cols"][k] for k, _ in columns] + [r["total"]])
            ws.append(["%s 소계" % g["date"], "", ""] + [g["sub"][k] for k, _ in columns] + [g["sub_total"]])
        ws.append(["총합계", "", ""] + [grand[k] for k, _ in columns] + [grand_total])
        for i, w in enumerate([12, 34, 8] + [11] * len(columns) + [10], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        fname = "작업집계_상품별_%s_%s.xlsx" % ("거점센터" if view == "hub" else "채널", today)
    else:
        ch = request.args.get("ch", "")
        if ch not in sc.CHANNELS:
            ch = CH_ORDER[0]
        blocks, grand, dates = _checklist(ch, sel_date)
        cfg = sc.CHANNELS[ch]
        ws.title = "검수리스트"
        qtylabel = "발주단위수량" if ch == "cu" else "수량"
        ws.append(["배송센터", "센터명", "상품명", "합계 : %s" % qtylabel])
        for c in range(1, 5):
            ws.cell(1, c).font = Font(name="맑은 고딕", size=11, bold=True)
            ws.cell(1, c).fill = PatternFill("solid", fgColor="FFDDEBF7"); ws.cell(1, c).border = border

        def put(r, vals, bold=False, fill=None):
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, v); cell.font = Font(name="맑은 고딕", size=10, bold=bold); cell.border = border
                if fill:
                    cell.fill = fill
        r = 2
        for b in blocks:
            first_hub = True
            for cb in b["centers"]:
                first_center = True
                for prod, q in cb["prods"]:
                    put(r, [b["hub"] if first_hub else "", cb["center"] if first_center else "", prod, q])
                    first_hub = first_center = False; r += 1
                put(r, ["", "%s 요약" % cb["center"], "", cb["sub"]], bold=True, fill=yellow); r += 1
        put(r, ["총합계", "", "", grand], bold=True, fill=yellow)
        for i, w in enumerate([20, 22, 30, 16], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        fname = "%s_검수리스트_%s.xlsx" % (cfg["name"], sel_date or "전체")

    ws.freeze_panes = "A2"
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------- 마스터 관리 ----------------
MASTER_DEFS = {
    "center":  {"file": sc.CENTER_MASTER, "title": "센터 (센터명→이고센터) · 전 채널 공통", "kind": "center"},
    "product": {"file": "product.json",   "title": "상품 (한익스상품명 ← 채널별 상품코드)", "kind": "product"},
}

@app.route("/masters")
def masters():
    counts = {k: len(sc.load_master(v["file"])) for k, v in MASTER_DEFS.items()}
    return render_template("masters.html", defs=MASTER_DEFS, counts=counts)


@app.route("/masters/<which>", methods=["GET"])
def master_edit(which):
    if which not in MASTER_DEFS:
        abort(404)
    d = MASTER_DEFS[which]
    data = sc.load_master(d["file"])
    q = request.args.get("q", "").strip()
    reps = []
    # 상품 마스터: 상품 1개 = 1행. 같은 상품이라도 채널마다 코드가 달라 코드는 한 칸에 모아 보여준다
    if d["kind"] == "product":
        items = sc.product_groups()
        reps = [g["rep"] for g in items]
        if q:
            ql = q.lower()
            items = [g for g in items
                     if ql in g["rep"].lower() or any(ql in c.lower() for c in g["codes"])]
    else:
        items = sorted(data.items())
        if q:
            items = [(k, v) for k, v in items if q.lower() in k.lower() or q.lower() in str(v).lower()]
    # 한 화면에 다 뿌리면 찾기 힘들어서 페이지로 나눈다
    per, found = 25, len(items)
    pages = max(1, (found + per - 1) // per)
    page = min(max(request.args.get("page", 1, type=int), 1), pages)
    items = items[(page - 1) * per:page * per]
    hubs = sc.all_hubs()
    return render_template("master_edit.html", which=which, d=d, items=items, reps=reps,
                           q=q, total=len(data), found=found, page=page, pages=pages,
                           hubs=hubs, product_cats=sc.PRODUCT_CATS)


@app.route("/masters/<which>/save", methods=["POST"])
def master_save(which):
    if which not in MASTER_DEFS:
        abort(404)
    d = MASTER_DEFS[which]
    data = sc.load_master(d["file"])
    is_prod = d["kind"] == "product"
    act = request.form.get("action")

    def save_product(old_rep):
        """상품 1건(이름·구분·코드들)을 통째로 저장. old_rep이 있으면 그 상품을 교체."""
        rep = request.form.get("value", "").strip()
        cat = request.form.get("cat", sc.DEFAULT_CAT).strip() or sc.DEFAULT_CAT
        codes = sc.split_codes(request.form.get("codes", ""))
        if not rep or not codes:
            flash("한익스상품명과 상품코드를 모두 입력해 주세요.", "error")
            return
        mine = {c for c in data if (sc.product_rep_cat(data, c)[0] or "") == old_rep} if old_rep else set()
        dup = [c for c in codes if c in data and c not in mine]
        if dup:
            flash("상품코드 %s 는 이미 다른 상품에 등록돼 있습니다." % ", ".join(dup), "error")
            return
        for c in mine:                       # 이 상품에서 빠진 코드는 제거
            data.pop(c)
        for c in codes:
            data[c] = {"rep": rep, "cat": cat}
        flash("저장: %s (%s) · 상품코드 %d개" % (rep, cat, len(codes)), "ok")

    if act == "add":
        if is_prod:
            save_product(None)
        else:
            k = request.form.get("key", "").strip()
            v = request.form.get("value", "").strip()
            if k and v:
                data[k] = v
                flash("추가: %s → %s" % (k, v), "ok")
    elif act == "update":
        if is_prod:
            save_product(request.form.get("old_rep", "").strip())
        else:
            k = request.form.get("key", "").strip()
            v = request.form.get("value", "").strip()
            if k in data and v:
                data[k] = v
                flash("수정: %s → %s" % (k, v), "ok")
    elif act == "delete":
        if is_prod:
            old = request.form.get("old_rep", "").strip()
            codes = [c for c in data if (sc.product_rep_cat(data, c)[0] or "") == old]
            for c in codes:
                data.pop(c)
            if codes:
                flash("삭제: %s (상품코드 %d개)" % (old, len(codes)), "ok")
        else:
            k = request.form.get("key", "").strip()
            if k in data:
                data.pop(k)
                flash("삭제: %s" % k, "ok")
    sc.save_master(d["file"], data)
    return redirect(url_for("master_edit", which=which, q=request.form.get("q", "")))


# ---------- 마스터 엑셀 (다운로드 → 엑셀에서 수정 → 업로드) ----------
MASTER_TMP = os.path.join(ARCHIVE_DIR, "master_upload")
MASTER_BAK = os.path.join(ARCHIVE_DIR, "master_backup")


@app.route("/masters/<which>/xlsx")
def master_xlsx(which):
    if which not in sc.MASTER_KINDS:
        abort(404)
    bio = sc.build_master_xlsx(which)
    today = datetime.datetime.now().strftime("%Y%m%d")
    nm = {"center": "센터마스터", "product": "상품마스터"}[which]
    return send_file(bio, as_attachment=True, download_name="%s_%s.xlsx" % (nm, today),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/masters/<which>/import", methods=["POST"])
def master_import(which):
    """엑셀 업로드 → 무엇이 바뀌는지 먼저 보여준다(아직 반영 안 함)."""
    if which not in sc.MASTER_KINDS:
        abort(404)
    f = request.files.get("file")
    if not f or not f.filename:
        flash("엑셀 파일을 선택해 주세요.", "error")
        return redirect(url_for("master_edit", which=which))
    if os.path.splitext(f.filename)[1].lower() not in (".xlsx", ".xlsm"):
        flash("xlsx 파일만 업로드할 수 있습니다.", "error")
        return redirect(url_for("master_edit", which=which))
    os.makedirs(MASTER_TMP, exist_ok=True)
    token = "%s_%s" % (which, uuid.uuid4().hex[:8])
    path = os.path.join(MASTER_TMP, token + ".xlsx")
    f.save(path)
    try:
        new, errors = sc.parse_master_xlsx(path, which)
    except Exception as e:
        os.remove(path)
        flash("엑셀을 읽지 못했습니다: %s" % e, "error")
        return redirect(url_for("master_edit", which=which))
    if not new and not errors:
        os.remove(path)
        flash("내용이 없는 엑셀입니다.", "error")
        return redirect(url_for("master_edit", which=which))
    diff = sc.diff_master(which, new)
    return render_template("master_import.html", which=which, d=MASTER_DEFS[which],
                           token=token, filename=f.filename, diff=diff,
                           errors=errors, total=len(new))


@app.route("/masters/<which>/import/apply", methods=["POST"])
def master_import_apply(which):
    """확인 후 실제 반영. 반영 직전 현재 마스터를 백업한다."""
    if which not in sc.MASTER_KINDS:
        abort(404)
    token = request.form.get("token", "")
    path = os.path.join(MASTER_TMP, token + ".xlsx")
    if not token.startswith(which + "_") or not os.path.exists(path):
        flash("업로드가 만료되었습니다. 엑셀을 다시 올려주세요.", "error")
        return redirect(url_for("master_edit", which=which))
    new, errors = sc.parse_master_xlsx(path, which)
    fname = sc.MASTER_KINDS[which]["file"]
    # 백업 후 교체
    os.makedirs(MASTER_BAK, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    with open(os.path.join(MASTER_BAK, "%s.%s.json" % (fname, ts)), "w", encoding="utf-8") as bf:
        json.dump(sc.load_master(fname), bf, ensure_ascii=False, indent=1, sort_keys=True)
    diff = sc.diff_master(which, new)
    sc.save_master(fname, new)
    os.remove(path)
    flash("엑셀을 반영했습니다 — 추가 %d · 수정 %d · 삭제 %d (전체 %d개). 이전 마스터는 백업해 두었습니다."
          % (len(diff["added"]), len(diff["changed"]), len(diff["removed"]), len(new)), "ok")
    return redirect(url_for("master_edit", which=which))


@app.route("/guide")
def guide():
    return render_template("guide.html", channels=CH_ORDER, cfg=sc.CHANNELS)


@app.route("/health")
def health():
    return jsonify(ok=True)


if __name__ == "__main__":
    import webbrowser, threading
    port = int(os.environ.get("PORT", 5057))
    host = os.environ.get("HOST", "127.0.0.1")   # 로컬 전용(방화벽 팝업 방지). 서버 배포 시 HOST=0.0.0.0
    if os.environ.get("NO_BROWSER") != "1":
        threading.Timer(1.2, lambda: webbrowser.open("http://127.0.0.1:%d/" % port)).start()
    app.run(host=host, port=port, debug=False)
